import openpyxl
from trio_websocket import open_websocket


def getRowNumber(filename, sheetname):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheetname]
    return Sheet.max_row


def readData(filename, sheetname, row_num, column_num):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheetname]
    return Sheet.cell(row=row_num, column=column_num).value


def writeData(filename, sheetname, row_num, column_num, data):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheetname]
    Sheet.cell(row=row_num, column=column_num).value = data
    Excel_File.save(filename)
